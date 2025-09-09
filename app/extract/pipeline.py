"""Email → Quote orchestration (Task D)."""

from __future__ import annotations

import os
from datetime import datetime
from pathlib import Path
from typing import Any

from sqlalchemy.orm import Session

from app.db.crud.emails import get_email_with_parts
from app.db.crud.quotes import get_or_create_quote, get_or_create_version, replace_items
from app.db.crud.vendors import upsert_vendor
from app.extract.llm import extract_quote_json
from app.extract.normalize import normalize_extraction
from app.extract.prefilter import likely_contains_quote


def _read_text_safe(path: Path, encoding: str = "utf-8") -> str | None:
    try:
        return path.read_text(encoding=encoding)
    except Exception:
        return None


def _extract_text_from_attachment(path: Path) -> str | None:
    suffix = path.suffix.lower()
    if suffix in {".txt", ""}:
        return _read_text_safe(path)
    if suffix == ".pdf":
        try:
            from pdfminer.high_level import extract_text

            return extract_text(str(path))
        except Exception:
            return None
    if suffix == ".docx":
        try:
            from docx import Document

            doc = Document(str(path))
            return "\n".join(p.text for p in doc.paragraphs)
        except Exception:
            return None
    if suffix in {".xlsx", ".xlsm"}:
        try:
            from openpyxl import load_workbook

            wb = load_workbook(filename=str(path), data_only=True, read_only=True)
            texts: list[str] = []
            for ws in wb.worksheets:
                for row in ws.iter_rows(values_only=True):
                    cells = [str(c) for c in row if c is not None]
                    if cells:
                        texts.append("\t".join(cells))
            return "\n".join(texts)
        except Exception:
            return None
    return None


def process_email(db: Session, email_id: int) -> dict[str, Any]:
    """Process a single email into quotes/versions/items. Idempotent by (quote_id, source_email_id)."""
    email = get_email_with_parts(db, email_id)
    if not email:
        return {"email_id": email_id, "processed": False, "reason": "email-not-found"}

    subject = email.subject or ""
    from_addr = email.from_addr
    to_addrs = email.to_addrs or []
    sent_at: datetime | None = email.sent_at
    date_str = sent_at.isoformat() if sent_at else None

    # Prefer text/plain body
    plain = next((b.body_text for b in email.bodies if b.mime_type == "text/plain" and b.body_text), None)
    html = next((b.body_html for b in email.bodies if b.mime_type == "text/html" and b.body_html), None)
    body_text = plain or (html or "")

    # Prefilter
    if not likely_contains_quote(subject, body_text):
        return {"email_id": email_id, "processed": False, "reason": "prefilter-skip"}

    # Attachment text
    attachments_texts: list[str] = []
    for a in email.attachments:
        if not a.local_path:
            continue
        p = Path(a.local_path)
        if not p.exists():
            continue
        text = _extract_text_from_attachment(p)
        if text:
            # limit size per attachment to avoid huge prompts
            attachments_texts.append(f"--- {p.name} ---\n" + text[:8000])
    attachments_text = "\n\n".join(attachments_texts) if attachments_texts else None

    # Call LLM
    extraction = extract_quote_json(
        subject=subject,
        from_=from_addr,
        to=to_addrs,
        date=date_str,
        body_text=body_text,
        attachments_text=attachments_text,
    )

    # Normalize
    extraction = normalize_extraction(extraction)

    # Persist vendor
    vendor_name = (extraction.get("vendor") or {}).get("name")
    vendor_domain = (extraction.get("vendor") or {}).get("domain")
    vendor_id = upsert_vendor(db, vendor_name, vendor_domain)

    # Create or get quote (thread + vendor)
    thread_id = email.thread_id
    quote = get_or_create_quote(db, thread_id=thread_id, vendor_id=vendor_id, anchor_email_id=email_id)

    # Create/update version for this email
    created_versions = 0
    for ver in extraction.get("versions", []):
        # Parse date if present
        valid_till_val = ver.get("valid_till")
        if isinstance(valid_till_val, str):
            try:
                from datetime import date as _date

                valid_till_parsed = _date.fromisoformat(valid_till_val)
            except Exception:
                valid_till_parsed = None
        else:
            valid_till_parsed = valid_till_val
        v = get_or_create_version(
            db,
            quote_id=quote.id,
            source_email_id=email_id,
            version_label=ver.get("version_label"),
            currency=ver.get("currency") or "",
            subtotal=ver.get("subtotal"),
            tax=ver.get("tax"),
            shipping=ver.get("shipping"),
            total=ver.get("total") or 0.0,
            valid_till=valid_till_parsed,
            terms=ver.get("terms"),
            extracted_json=ver,
        )
        replace_items(db, v.id, ver.get("items", []))
        created_versions += 1

    db.commit()
    return {"email_id": email_id, "processed": True, "versions": created_versions}
